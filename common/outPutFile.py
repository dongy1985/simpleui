import shutil
import time, datetime
import openpyxl
import os
from django.contrib import admin, messages
from django.db import transaction
from django.shortcuts import render
from django.contrib.admin import actions
from django.urls import reverse
from django.contrib import admin
from django.utils.encoding import escape_uri_path
from django.db.models import Q
from import_export.admin import ImportExportModelAdmin, ImportExportActionModelAdmin
from common.models import *
from company.models import *
from common.const import const
from copy import copy
from aggregation.models import *


def export(folder_name, datFrom, datTo):
    datFrom_Y = datFrom[0:4]
    datFrom_M = datFrom[5:7]
    datTo_Y = datTo[0:4]
    datTo_M = datTo[5:7]
    # 月数
    counts = (int(datTo_Y) - int(datFrom_Y)) * 12 + (int(datTo_M) - int(datFrom_M))
    fileName = copyExl(folder_name, const.SHEET_YEAR)
    # mkExcel
    mkExcel(fileName, counts, datFrom, datTo)


def copyExl(folder_name, userName):
    # コピー先のパス設定
    fileName = const.DIR + folder_name + const.FILESTART + userName + const.XLS
    # コピー
    shutil.copyfile(const.YEAR_TEMPLATEPATH, fileName)
    return fileName


# 年度単位の集計表dataの作成
def mkExcel(fileName, counts, datFrom, datTo):
    # コピー先ファイル名の取得
    book = openpyxl.load_workbook(fileName)
    # コピー先シート名の取得
    sheet = book.get_sheet_by_name(const.SHEET_YEAR)

    #ヘーダエリア(１～３行目)
    count = 1
    while count <= counts:
        for i in range(3, 9):
            # 月分の列を定義
            e1 = sheet.cell(row=1, column=i + 6 * count)
            sheet[e1.coordinate]._style = sheet.cell(row=1, column=i)._style
            # コピー元の列を定義
            copy = sheet.cell(row=2, column=i).value
            sheet.cell(row=2, column=i + 6 * count, value=copy)
            e2 = sheet.cell(row=2, column=i + 6 * count)
            sheet[e2.coordinate]._style = sheet.cell(row=2, column=i)._style
            e3 = sheet.cell(row=3, column=i + 6 * count)
            sheet[e3.coordinate]._style = sheet.cell(row=3, column=i)._style
        sheet.merge_cells(const.SERU[count])
        count += 1

    datFrom_Y = int(datFrom[0:4])
    datFrom_M = int(datFrom[5:7])
    count = 0
    # data write
    while count <= counts:
        if datFrom_M <= 12:
            if datFrom_M < 10:
                temp_queryset = Aggregation.objects.filter(
                    Q(attendance_YM__startswith=(str(datFrom_Y) + '-0' + str(datFrom_M)))
                )
            else:
                temp_queryset = Aggregation.objects.filter(
                    Q(attendance_YM__startswith=(str(datFrom_Y) + '-' + str(datFrom_M)))
                )
            datFrom_M = datFrom_M + 1
        else:
            datFrom_Y = datFrom_Y + 1
            temp_queryset = Aggregation.objects.filter(
                Q(attendance_YM__startswith=(str(datFrom_Y) + '-01'))
            )
            datFrom_M = 2

        # 年度単位の集計表ヘッダー部の取得
        headMst = CrdMst.objects.filter(tplType=const.AGG_YEAR, crdDiv=const.CRD_DIV_H, delFlg=const.DEL_FLG_0). \
            values_list('crdY', 'crdX', 'defVal').order_by('itemSort')
        # 年度単位の集計表明細部の取得
        dataMst = CrdMst.objects.filter(tplType=const.AGG_YEAR, crdDiv=const.CRD_DIV_D, delFlg=const.DEL_FLG_0). \
            values_list('crdY', 'crdX', 'defVal').order_by('itemSort')
        data_row = dataMst[0][0]
        j = 0
        for obj in temp_queryset:
            data_row = data_row + 1
            # 社員番号
            aggEmpNo = obj.empNo
            # 社員名前
            aggName = obj.name
            # 実働時間
            aggWorkTime = obj.working_time
            # 出勤日数
            aggAttendCount = obj.attendance_count
            # 欠勤日数
            aggAbsCount = obj.absence_count
            # 年休日数
            aggAnnLeave = obj.annual_leave
            # 休出日数
            aggRestCount = obj.rest_count
            # 遅早退日数
            aggLateCount = obj.late_count
            # 統計年月
            objMonth = obj.attendance_YM.strftime('%Y-%m')

            # write data
            # 報告月
            sheet.cell(1, 3 + count * 6, objMonth)
            if count == 0:
                # 社員番号
                sheet.cell(data_row, 1 + count * 6, aggEmpNo)
                # 氏名
                sheet.cell(data_row, 2 + count * 6, aggName)
            # 実働時間
            sheet.cell(data_row, 3 + count * 6, aggWorkTime)
            # 出勤
            sheet.cell(data_row, 4 + count * 6, aggAttendCount)
            # 欠勤
            sheet.cell(data_row, 5 + count * 6, aggAbsCount)
            # 年休
            sheet.cell(data_row, 6 + count * 6, aggAnnLeave)
            # 休出
            sheet.cell(data_row, 7 + count * 6, aggRestCount)
            # 遅早退
            sheet.cell(data_row, 8 + count * 6, aggLateCount)
            j = j+1
            # 行のコピー
            for a in range(1, 100):
                e = sheet.cell(row=j + 1, column=a)
                sheet[e.coordinate]._style = sheet.cell(row=3, column=a)._style
        count += 1
    # save
    book.save(fileName)