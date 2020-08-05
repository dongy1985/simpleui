import time, datetime
import shutil
import openpyxl
import os
import math

from django.contrib import admin, messages
from django.db import transaction
from django.shortcuts import render
from django.contrib.admin import actions
from django.urls import reverse
from django.contrib import admin
from django.utils.encoding import escape_uri_path
from django.db.models import Q

from attendance.models import *
from import_export import resources
from import_export.admin import ImportExportModelAdmin, ImportExportActionModelAdmin
from common.models import *
from company.models import *
from common.const import const


def export(queryset, folder_name):
    # temp id
    tempId = ''
    # tempYM
    tempYM = ''
    temp_queryset = None
    # querysetFilter
    for obj in queryset:
        workYM = obj.date.strftime('%Y-%m')
        if obj.user_id != tempId or workYM != tempYM:
            temp_queryset = queryset.filter(
                Q(user_id=obj.user_id)
                & Q(date__startswith=workYM)
            )
            # call make Excel
            mkExcel(temp_queryset, folder_name)
            # 更新临时数据
            tempId = obj.user_id
            tempYM = workYM
            queryset = queryset.filter(
                ~(Q(user_id=obj.user_id)
                  & Q(date__startswith=workYM))
            )
            return export(queryset=queryset, folder_name=folder_name)


def copyExl(folder_name, userName, objYear, objMonth):
    fileName = const.DIR + folder_name + const.FILESTART + userName + const.UNDERLINE + str(objYear) + str(
        objMonth) + const.XLSM
    shutil.copyfile(const.TEMPLATEPATH, fileName)
    return fileName


def mkExcel(queryset, folder_name):
    # get data
    for obj in queryset:
        userName = obj.name
        objMonth = obj.date.month
        objYear = obj.date.year
    # copy
    fileName = copyExl(folder_name, userName, objYear, objMonth)
    # book = openpyxl.load_workbook(fileName)
    book = openpyxl.load_workbook(fileName, keep_vba=True)
    sheet = book.get_sheet_by_name(const.SHEET)

    # head data write
    headMst = CrdMst.objects.filter(tplType=const.TPL_XLS, crdDiv=const.CRD_DIV_H, delFlg=const.DEL_FLG_0). \
        values_list('crdY', 'crdX', 'defVal').order_by('itemSort')
    for obj in queryset:
        userNumber = Employee.objects.get(user=obj.user_id).empNo
        userName = obj.name
        objMonth = obj.date.month
        objYear = obj.date.year
        projectDate = obj.date
        userId = Employee.objects.get(user=obj.user_id).empNo
        # 現場
        if len(WorkSiteDetail.objects.filter(
                Q(member_id=userId)
                & Q(from_date__lte=projectDate)
                & Q(to_date__gte=projectDate)
            ).values('manager_id')) != 0:
            temp_Site_id = WorkSiteDetail.objects.filter(
                Q(member_id=userId)
                & Q(from_date__lte=projectDate)
                & Q(to_date__gte=projectDate)
            ).values_list('manager_id')
            site_id = temp_Site_id[0][0]
            siteNumber = WorkSite.objects.get(id=site_id).site_number
            projectName = WorkSite.objects.get(id=site_id).project_name
            managerId = WorkSite.objects.get(id=site_id).manager_id
            managerName = Employee.objects.get(empNo=managerId).name
    # 社員番号
    sheet.cell(headMst[0][0], headMst[0][1], userNumber)
    # 氏名
    sheet.cell(headMst[1][0], headMst[1][1], userName)
    # 報告月
    sheet.cell(headMst[2][0], headMst[2][1], str(objYear) + '/' + str(objMonth) + '/1')
    # 発注番号
    sheet.cell(headMst[3][0], headMst[3][1], siteNumber)
    # 案件名
    sheet.cell(headMst[4][0], headMst[4][1], projectName)
    # 責任者名
    sheet.cell(headMst[5][0], headMst[5][1], managerName)

    # duty data write
    dataMst = CrdMst.objects.filter(tplType=const.TPL_XLS, crdDiv=const.CRD_DIV_D, delFlg=const.DEL_FLG_0). \
        values_list('crdY', 'crdX', 'defVal').order_by('itemSort')
    data_row = dataMst[0][0]
    for obj in queryset:
        # start_time
        sheet.cell(data_row + obj.date.day, dataMst[0][1], obj.start_time)
        # end_time
        sheet.cell(data_row + obj.date.day, dataMst[1][1], obj.end_time)
        # restTime
        # 0.0 -> 00:00
        temp_modf = math.modf(float(obj.rest))
        temp_hour = int(temp_modf[1])
        temp_minute = int(temp_modf[0] * 60)
        temp_rest_time = str(temp_hour) + ':' + str(temp_minute)
        sheet.cell(data_row + obj.date.day, dataMst[2][1], temp_rest_time)
        # contents
        sheet.cell(data_row + obj.date.day, dataMst[3][1], obj.contents)
    # save
    book.save(fileName)


# 月度単位の集計表(excel)の導出


def exportExcel(folder_name, datFrom):
    # DBから該当月度のデータ取得
    temp_queryset = DutyStatistics.objects.filter(
        Q(attendance_YM__startswith=datFrom)
    )
    # 月度単位の集計表templateの作成
    fileName = copyExle(folder_name, datFrom)
    # 【月度単位の集計表dataの作成】を呼び出し
    mkExl(temp_queryset, fileName, datFrom)
    return fileName


# 月度単位の集計表templateの作成
def copyExle(folder_name, datFrom):
    # コピー先のパス設定
    fileName = const.DIR + folder_name + const.FILESTART + const.SHEET_MONTH + const.UNDERLINE + str(
        datFrom) + const.XLS
    # コピー
    shutil.copyfile(const.MONTH_TEMPLATEPATH, fileName)
    return fileName


# 月度単位の集計表dataの作成
def mkExl(queryset, fileName, datFrom):
    # コピー先ファイル名の取得
    book = openpyxl.load_workbook(fileName)
    # コピー先シート名の取得
    sheet = book.get_sheet_by_name(const.SHEET_MONTH)

    # 月度単位の集計表ヘッダー部の取得
    headMst = CrdMst.objects.filter(tplType=const.AGG_MONTH, crdDiv=const.CRD_DIV_H, delFlg=const.DEL_FLG_0). \
        values_list('crdY', 'crdX', 'defVal').order_by('itemSort')
    # 月度単位の集計表明細部の取得
    dataMst = CrdMst.objects.filter(tplType=const.AGG_MONTH, crdDiv=const.CRD_DIV_D, delFlg=const.DEL_FLG_0). \
        values_list('crdY', 'crdX', 'defVal').order_by('itemSort')
    data_row = dataMst[0][0]
    i = 3
    for obj in queryset:
        data_row = data_row + 1
        # 社員番号
        aggEmpNo = obj.empNo
        # 社員名前
        aggName = obj.name
        # 実働時間
        aggWorkTime = obj.working_time
        # 出勤日数
        aggAttendCount = obj.attendance_count
        # 欠勤日数
        aggAbsCount = obj.absence_count
        # 年休日数
        aggAnnLeave = obj.annual_leave
        # 休出日数
        aggRestCount = obj.rest_count
        # 遅早退日数
        aggLateCount = obj.late_count

        # 社員番号
        sheet.cell(data_row, dataMst[0][1], aggEmpNo)
        # 社員名前
        sheet.cell(data_row, dataMst[1][1], aggName)
        # 実働時間
        sheet.cell(data_row, dataMst[2][1], aggWorkTime)
        # 出勤日数
        sheet.cell(data_row, dataMst[3][1], aggAttendCount)
        # 欠勤日数
        sheet.cell(data_row, dataMst[4][1], aggAbsCount)
        # 年休日数
        sheet.cell(data_row, dataMst[5][1], aggAnnLeave)
        # 休出日数
        sheet.cell(data_row, dataMst[6][1], aggRestCount)
        # 遅早退日数
        sheet.cell(data_row, dataMst[7][1], aggLateCount)
        # 統計年月
        sheet.cell(headMst[0][0], headMst[0][1], str(datFrom))
        # 行ごとに書式のコピー
        i = i + 1
        for a in range(1, 9):
            nextrow = sheet.cell(row=i + 1, column=a)
            sheet[nextrow.coordinate]._style = sheet.cell(row=5, column=a)._style

    book.save(fileName)


# 年度単位の集計表(excel)の導出
def exportYearExcel(folder_name, attendance_YM_From, attendance_YM_To, queryset):
    # 統計年月開始年
    attendance_YM_From_Y = attendance_YM_From[0:4]
    # 統計年月開始月
    attendance_YM_From_M = attendance_YM_From[5:7]
    # 統計年月終了年
    attendance_YM_To_Y = attendance_YM_To[0:4]
    # 統計年月終了月
    attendance_YM_To_M = attendance_YM_To[5:7]
    # 統計月数：統計年月開始から、統計年月終了まで、何か月
    counts = (int(attendance_YM_To_Y) - int(attendance_YM_From_Y)) * 12 + (
                int(attendance_YM_To_M) - int(attendance_YM_From_M))
    # ファイルの名
    fileName = copyExcel(folder_name, const.SHEET_YEAR)
    # 統計期間のユーザー数
    user_name = []
    user_Con = []
    for obj in queryset:
        user_name.append(obj.name)
    for user in user_name:
        if user not in user_Con:
            user_Con.append(user)
    # 【年度単位の集計表dataの作成】を呼び出し
    makeExcel(fileName, counts, attendance_YM_From, user_Con)
    return fileName


# 年度単位の集計表templateの作成
def copyExcel(folder_name, userName):
    # コピー先のパス設定
    fileName = const.DIR + folder_name + const.FILESTART + userName + const.XLS
    # コピー
    shutil.copyfile(const.YEAR_TEMPLATEPATH, fileName)
    return fileName


# 年度単位の集計表dataの作成
def makeExcel(fileName, counts, attendance_YM_From, user_Con):
    # コピー先ファイル名の取得
    book = openpyxl.load_workbook(fileName)
    # コピー先シート名の取得
    sheet = book.get_sheet_by_name(const.SHEET_YEAR)

    # ヘーダエリア(１～３行目)
    count = 1
    while count <= counts:
        for i in range(3, 9):
            # 月分の列を定義
            # ヘーダエリア(１行目)
            e1 = sheet.cell(row=1, column=i + 6 * count)
            sheet[e1.coordinate]._style = sheet.cell(row=1, column=i)._style
            # コピー元の列を定義
            # ヘーダエリア(2行目):valueのコピー
            copy = sheet.cell(row=2, column=i).value
            sheet.cell(row=2, column=i + 6 * count, value=copy)
            # ヘーダエリア(2行目):セルのコピー
            e2 = sheet.cell(row=2, column=i + 6 * count)
            sheet[e2.coordinate]._style = sheet.cell(row=2, column=i)._style
            # ヘーダエリア(3行目):セルのコピー
            e3 = sheet.cell(row=3, column=i + 6 * count)
            sheet[e3.coordinate]._style = sheet.cell(row=3, column=i)._style
        # セル結合
        sheet.merge_cells(const.SERU[count])
        count += 1
    # 行のコピー
    for j in range(1, len(user_Con)):
        for a in range(1, 3 + (counts + 1) * 6):
            e = sheet.cell(row=j + 3, column=a)
            sheet[e.coordinate]._style = sheet.cell(row=3, column=a)._style

    del_column = []
    for con in range(len(user_Con)):
        # 統計年月開始年
        attendance_YM_From_Y = int(attendance_YM_From[0:4])
        # 統計年月開始月
        attendance_YM_From_M = int(attendance_YM_From[5:7])
        count = 0
        while count <= counts:
            if attendance_YM_From_M <= 12:
                if attendance_YM_From_M < 10:
                    temp_queryset = DutyStatistics.objects.filter(
                        Q(attendance_YM__startswith=(str(attendance_YM_From_Y) + '-0' + str(attendance_YM_From_M)))
                        & Q(name=str(user_Con[con]))
                    )
                else:
                    temp_queryset = DutyStatistics.objects.filter(
                        Q(attendance_YM__startswith=(str(attendance_YM_From_Y) + '-' + str(attendance_YM_From_M)))
                        & Q(name=str(user_Con[con]))
                    )
                attendance_YM_From_M = attendance_YM_From_M + 1
            else:
                attendance_YM_From_Y = attendance_YM_From_Y + 1
                temp_queryset = DutyStatistics.objects.filter(
                    Q(attendance_YM__startswith=(str(attendance_YM_From_Y) + '-01'))
                    & Q(name=str(user_Con[con]))
                )
                attendance_YM_From_M = 2
            # write data
            for obj in temp_queryset:
                # 社員番号
                aggEmpNo = obj.empNo
                # 社員名前
                aggName = obj.name
                # 実働時間
                aggWorkTime = obj.working_time
                # 出勤日数
                aggAttendCount = obj.attendance_count
                # 欠勤日数
                aggAbsCount = obj.absence_count
                # 年休日数
                aggAnnLeave = obj.annual_leave
                # 休出日数
                aggRestCount = obj.rest_count
                # 遅早退日数
                aggLateCount = obj.late_count
                # 統計年月
                objMonth = obj.attendance_YM.strftime('%Y-%m')

                # write data
                # 報告月
                sheet.cell(1, 3 + count * 6, objMonth)
                # 社員番号
                sheet.cell(con + 3, 1, aggEmpNo)
                # 氏名
                sheet.cell(con + 3, 2, aggName)
                # 実働時間
                sheet.cell(con + 3, 3 + count * 6, aggWorkTime)
                # 出勤
                sheet.cell(con + 3, 4 + count * 6, aggAttendCount)
                # 欠勤
                sheet.cell(con + 3, 5 + count * 6, aggAbsCount)
                # 年休
                sheet.cell(con + 3, 6 + count * 6, aggAnnLeave)
                # 休出
                sheet.cell(con + 3, 7 + count * 6, aggRestCount)
                # 遅早退
                sheet.cell(con + 3, 8 + count * 6, aggLateCount)
            count += 1
    # データがないのセルを削除する
    # 削除行列
    del_cell = []
    # 削除開始の列:3
    column = 3
    # 削除開始< 全部結果行列
    while column <= 3 + counts*6:
        # 削除フラグ
        del_flg = True
        # ターゲット列：3, 3 + len(user_Con)
        for row in range(3, 3 + len(user_Con)):
            # セルの内容はあるの時は、削除フラグはFalse
            if sheet.cell(row=row, column=column).value != None:
                del_flg = False
                break
        # 削除フラグはTrueの時、ターゲット削除列を追加
        if del_flg:
            del_cell.append(column)
        # 6列を循环
        column += 6
    # 削除行列を循环：6列を循环を削除する
    for cell in del_cell:
        for i in range(6):
            sheet.delete_cols(cell)
    # save
    book.save(fileName)
