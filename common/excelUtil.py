import time, datetime
import shutil
import openpyxl
import os

from django.contrib import admin, messages
from django.db import transaction
from django.shortcuts import render
from django.contrib.admin import actions
from django.urls import reverse
from django.contrib import admin
from django.utils.encoding import escape_uri_path
from django.db.models import Q

from attendance.models import *
from aggregation.models import *
from import_export import resources
from import_export.admin import ImportExportModelAdmin, ImportExportActionModelAdmin
from common.models import *
from company.models import *
from common.const import const

# 月度単位の集計表(excel)の導出


def exportExcel(queryset, folder_name):
    # 月度の取得（検索期間として設定）
    for obj in queryset:
        workYM = obj.attendance_YM.strftime('%Y-%m')
        # DBから該当月度のデータ取得
        temp_queryset = Aggregation.objects.filter(
            Q(attendance_YM__startswith=workYM)
        )
        # 【月度単位の集計表dataの作成】を呼び出し
        mkExcel(temp_queryset, folder_name)

        queryset = queryset.filter(
            ~(Q(attendance_YM__startswith=workYM))
        )

        return exportExcel(queryset=queryset, folder_name=folder_name)

# 月度単位の集計表templateの作成
def copyExl(folder_name, objMonth):
    # コピー先のパス設定
    fileName = const.DIR + folder_name + const.FILESTART + const.SHEET_MONTH + const.UNDERLINE + str(objMonth) + const.XLS
    # コピー
    shutil.copyfile(const.MONTH_TEMPLATEPATH, fileName)
    return fileName

# 月度単位の集計表dataの作成
def mkExcel(queryset, folder_name):
    # 月度の取得
    for obj in queryset:
        objMonth = obj.attendance_YM.strftime('%Y-%m')
    # 月度単位の集計表templateの作成
    fileName = copyExl(folder_name, objMonth)
    # コピー先ファイル名の取得
    book = openpyxl.load_workbook(fileName)
    # コピー先シート名の取得
    sheet = book.get_sheet_by_name(const.SHEET_MONTH)

    # 月度単位の集計表ヘッダー部の取得
    headMst = CrdMst.objects.filter(tplType=const.AGG_MONTH, crdDiv=const.CRD_DIV_H, delFlg=const.DEL_FLG_0).\
    values_list('crdY', 'crdX', 'defVal').order_by('itemSort')
    # 月度単位の集計表明細部の取得
    dataMst = CrdMst.objects.filter(tplType=const.AGG_MONTH, crdDiv=const.CRD_DIV_D, delFlg=const.DEL_FLG_0).\
        values_list('crdY', 'crdX', 'defVal').order_by('itemSort')
    data_row = dataMst[0][0]
    i = 3
    for obj in queryset:
        data_row = data_row + 1
        # 社員番号
        aggEmpNo = obj.empNo
        # 社員名前
        aggName = obj.name
        # 実働時間
        aggWorkTime = obj.working_time
        # 出勤日数
        aggAttendCount = obj.attendance_count
        # 欠勤日数
        aggAbsCount = obj.absence_count
        # 年休日数
        aggAnnLeave = obj.annual_leave
        # 休出日数
        aggRestCount = obj.rest_count
        # 遅早退日数
        aggLateCount = obj.late_count
        # 統計年月
        objMonth = obj.attendance_YM.strftime('%Y-%m')

        # 社員番号
        sheet.cell(data_row, dataMst[0][1], aggEmpNo)
        # 社員名前
        sheet.cell(data_row, dataMst[1][1], aggName)
        # 実働時間
        sheet.cell(data_row, dataMst[2][1], aggWorkTime)
        # 出勤日数
        sheet.cell(data_row, dataMst[3][1], aggAttendCount)
        # 欠勤日数
        sheet.cell(data_row, dataMst[4][1], aggAbsCount)
        # 年休日数
        sheet.cell(data_row, dataMst[5][1], aggAnnLeave)
        # 休出日数
        sheet.cell(data_row, dataMst[6][1], aggRestCount)
        # 遅早退日数
        sheet.cell(data_row, dataMst[7][1], aggLateCount)
        # 統計年月
        sheet.cell(headMst[0][0], headMst[0][1], str(objMonth))
        # 行ごとに書式のコピー
        i = i + 1
        for a in range(1, 9):
            nextrow = sheet.cell(row=i + 1, column=a)
            sheet[nextrow.coordinate]._style = sheet.cell(row=5, column=a)._style

    book.save(fileName)

